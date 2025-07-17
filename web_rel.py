import json
from pathlib import Path
from time import sleep
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import Select
from selenium.webdriver.support.wait import WebDriverWait

ROOT_FOLDER = Path(__file__).parent
CHROME_DRIVER_EXE = ROOT_FOLDER / "chromedriver-win64" / "chromedriver.exe"
CONFIG = ROOT_FOLDER / "config.json"
NAV_KEY = ROOT_FOLDER / "nav_key.json"


def read_json(path_arg=""):
    with open(path_arg, 'r') as file:
        json_data = json.load(file)
        return json_data


def make_browser() -> webdriver.Chrome:
    options = webdriver.ChromeOptions()
    services = Service(executable_path=CHROME_DRIVER_EXE)
    browser = webdriver.Chrome(options=options, service=services)
    return browser


def make_click(arg_local=""):
    local_click = nav_elements[arg_local]
    element_clickable = EC.element_to_be_clickable
    clickable = WebDriverWait(browser, 12.0).until(
        element_clickable(
            (By.XPATH, local_click)
        )
    )
    return clickable


if not CONFIG.exists():
    input_user = input("Insert your user name (name.lastname): ")
    input_password = input("Insert your password: ")
    print("WARNING: don't insert the file, only the path to folder!")
    xlsx_local = input("Insert the path to file xlsx: ")

    dict_dump = {
        "nome-usuario": input_user,
        "password": input_password
    }

    with open(CONFIG, 'w') as file:
        json.dump((dict_dump, xlsx_local), file)


login_elements, xlsx_path = read_json(CONFIG)
nav_elements = read_json(NAV_KEY)

browser = make_browser()
browser.get(nav_elements["site"])

sleep(10)

for search_id, data_user in login_elements.items():
    login = make_click(search_id)
    login.click()
    login.send_keys(data_user)
    login.send_keys(Keys.ENTER)
sleep(5)

perfil = Select(
    WebDriverWait(browser, 12.5).until(
        EC.presence_of_element_located(
            (By.TAG_NAME, "select")
        )
    )
)
selection = browser.find_element(By.XPATH, "//button[contains(text(), 'Selecionar')]")
perfil.select_by_visible_text("PRD-SBT-Gestor")
selection.send_keys(Keys.ENTER)
sleep(5)

for click_operation in nav_elements:
    if click_operation in ["site", "nome-usuario", "password"]:
        continue
    navegation = make_click(click_operation)
    navegation.click()
    sleep(5)
browser.quit()


import csv
from datetime import date
from openpyxl import load_workbook

csv_file_name = f"LISTAGEM DAS ANOTAÇÕES_{date.today().strftime("%d-%m-%Y")}.csv"
csv_file = Path.home() / "Download" / csv_file_name

xlsx_p = Path(xlsx_path)
xlsx_file = xlsx_p / "CONTROLE_ANOTAÇÕES_2025.xlsx"

with open(csv_file, 'r', encoding="utf-8") as file:
    data = list(csv.reader(file))

workbook = load_workbook(str(xlsx_file))
table = workbook["L.A"]
table.delete_rows(1, table.max_row)

for index_line, line in enumerate(data, start=1):
    for index_column, data_value in enumerate(index_line, start=1):
        table.cell(row=index_line, column=index_column, value=data_value)

workbook.save()
